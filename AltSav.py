import sys
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from PIL import ImageGrab
import pyreadstat
import openpyxl

# Calculate Maximum Screen Width and Height
img = ImageGrab.grab()
w, h = img.size[0], img.size[1] - 80
boolExport = False

# Creating Main Window
class Main(QMainWindow):
    def __init__(self):
        global w, h
        super().__init__()
        self.setWindowTitle('AltSav')
        self.setGeometry(0, 0, w, h)
        self.setFixedSize(self.size())

        self.UI()
        self.show()

    def UI(self):
        self.menuUI()
        self.tabWidget()
        self.widgets()
        self.layouts()

    def menuUI(self):

        # Overall Menu
        self.menu = self.menuBar()
        self.file = self.menu.addMenu('&File')
        self.about = self.menu.addMenu('&About')

        # File Options
        # Open File
        self.open = QAction('Open')
        self.file.addAction(self.open)
        self.open.triggered.connect(self.funcOpenSav)

        # Export File
        self.export = QAction('Export')
        self.file.addAction(self.export)
        self.export.setVisible(False)
        self.export.triggered.connect(self.funcExport)

        # Exit File
        self.exit = QAction('Exit')
        self.file.addAction(self.exit)
        self.exit.triggered.connect(self.funcExit)

        # About Options
        self.info = QAction('Info')
        self.about.addAction(self.info)
        self.info.triggered.connect(self.funcAbout)

    def tabWidget(self):
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        # Variable View
        self.tab1 = QWidget()
        self.tabs.addTab(self.tab1, 'Variable View')

    def widgets(self):
        # Tab 1 Widgets
        # Top Widget
        self.filterEntry = QLineEdit()
        self.filterEntry.setPlaceholderText('Enter variable name')
        self.searchBtn = QPushButton('Search')
        self.filterEntry.returnPressed.connect(self.searchTable)
        self.searchBtn.clicked.connect(self.searchTable)

        # Bottom Left Widget
        self.schemaTable = QTableWidget()
        self.schemaTable.setColumnCount(3)
        self.schemaTable.setHorizontalHeaderItem(0, QTableWidgetItem('Variable'))
        self.schemaTable.setHorizontalHeaderItem(1, QTableWidgetItem('Variable Label'))
        self.schemaTable.setHorizontalHeaderItem(2, QTableWidgetItem('Type'))
        self.schemaTable.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.schemaTable.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.schemaTable.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.schemaTable.verticalHeader().hide()
        self.schemaTable.itemSelectionChanged.connect(self.displayStats)

        # Bottom Right Widget
        self.freqTable = QTableWidget()
        self.freqTable.setColumnCount(4)
        self.freqTable.setHorizontalHeaderItem(0, QTableWidgetItem('Value'))
        self.freqTable.setHorizontalHeaderItem(1, QTableWidgetItem('Label'))
        self.freqTable.setHorizontalHeaderItem(2, QTableWidgetItem('n'))
        self.freqTable.setHorizontalHeaderItem(3, QTableWidgetItem('%'))
        self.freqTable.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.freqTable.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.freqTable.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.freqTable.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.freqTable.verticalHeader().hide()

    def layouts(self):
        # Tab 1 Layout
        # Creating Layouts
        self.mainLayout = QVBoxLayout()
        self.topLayout = QHBoxLayout()
        self.bottomLayout = QHBoxLayout()
        self.bottomLeftLayout = QVBoxLayout()
        self.bottomRightLayout = QVBoxLayout()

        # Adding Widgets to Top Layout
        self.topLayout.addWidget(self.filterEntry)
        self.topLayout.addWidget(self.searchBtn)
        for i in range(7):
            self.topLayout.addStretch()

        # Adding Widgets to bottom Layouts
        self.bottomLeftLayout.addWidget(self.schemaTable)
        self.bottomRightLayout.addWidget(self.freqTable)
        self.bottomLayout.addLayout(self.bottomLeftLayout, 70)
        self.bottomLayout.addLayout(self.bottomRightLayout, 30)

        # Adding Layouts to MainLayout
        self.mainLayout.addLayout(self.topLayout)
        self.mainLayout.addLayout(self.bottomLayout)
        self.tab1.setLayout(self.mainLayout)

    def funcOpenSav(self):

        global df, dfText, metaVarToLabels, metaVarToFormat, metaVarToValueLabels, boolExport

        try:
            # Capturing filename
            directory = QFileDialog.getOpenFileName(self, 'Open Sav File', '', 'Sav Files (*.sav)')
            fileName = directory[0]

            if fileName != '':

                # Clear all data from Schema Table and Frequency Table
                for i in reversed(range(self.schemaTable.rowCount())):
                    self.schemaTable.removeRow(i)
                for i in reversed(range(self.freqTable.rowCount())):
                    self.freqTable.removeRow(i)
                qApp.processEvents()

                try:
                    # Read Sav File and storing dataframe and meta data using regular approach
                    df, meta = pyreadstat.read_sav(fileName)
                    dfText = pyreadstat.set_value_labels(df, meta, formats_as_category=True)

                except UnicodeDecodeError:

                    # In case of error we try to identify the error variables and try to load the data atleast for other variables
                    QMessageBox.information(self, 'Info', 'It seems that some of your string variables have different '
                                                          'encoding format (not readable by UTF-8).\n'
                                                          'Program will now try to identify these variables and read file '
                                                          'excluding these variables')

                    # Read only metadata and check for error amongst all string variables
                    df, meta = pyreadstat.read_sav(fileName, metadataonly=True)
                    listTotalVar = meta.column_names
                    listStringVar = [x for x, y in dict(meta.original_variable_types).items() if y[0] == 'A']
                    listErrorVar = []
                    for var in listStringVar:
                        try:
                            df, meta = pyreadstat.read_sav(fileName, usecols=[str(var)])
                        except:
                            listErrorVar.append(var)
                    listErrorVar =set(listErrorVar)

                    # Final Variable that needs to be loaded is TotalVariable - ErrorVariables
                    listFinal = [item for item in listTotalVar if item not in listErrorVar]
                    listFinal = list(set(listFinal))

                    # Read Sav File and storing dataframe and meta data for selected variables
                    df, meta = pyreadstat.read_sav(fileName, usecols=listFinal)
                    dfText = pyreadstat.set_value_labels(df, meta, formats_as_category=True)

                    # Pop Up Message to inform about skipped variable
                    mbox = QMessageBox()
                    mbox.setIcon(QMessageBox.Information)
                    message = ('Below variables are not read :- \n')
                    for var in listErrorVar:
                        message = message + str(var) + '\n'
                    mbox.setText(message)
                    mbox.setWindowTitle('Info')
                    mbox.setStandardButtons(QMessageBox.Ok)
                    mbox.exec_()

                # Dictionary Variables from Metadata
                metaVarToLabels = meta.column_names_to_labels
                metaVarToFormat = meta.original_variable_types
                metaVarToValueLabels = meta.variable_value_labels

                # Displaying Output for schema Table
                self.displaySchema()

                # Set Export Option On
                if boolExport == False:
                    self.export.setVisible(True)
                    boolExport = True

        except:
            QMessageBox.information(self, 'Info', 'Oops something went wrong while reading file.')

    def funcExport(self):
        # Setup new window for export options
        self.exportWindow = Export()

    def funcExit(self):
        mbox = QMessageBox.question(self, 'Info', 'Are you sure to exit?', QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if mbox == QMessageBox.Yes:
            sys.exit()

    def funcAbout(self):

        # About/Info Page

        mbox = QMessageBox()
        mbox.setIcon(QMessageBox.Information)
        mbox.setTextFormat(Qt.RichText)
        message = ("""
        Version - 0.1<br><br>
        This application was created by <a href='www.linkedin.com/in/shridharsahu19'>Shridhar Sahu</a><br><br>
        Features -<br>
        1. Reads .sav files<br>
        2. Displays variable name, label and type with search box<br>
        3. Displays frequency and stats on selection<br>
        4. Export data to excel as value or text along with an optional datamap<br><br>
        In case of any bugs or features required please reach out either via
        <a href='www.linkedin.com/in/shridharsahu19'>Linkedin</a> or
        <a href='https://forms.gle/KqMF9HoB8wu7pBHE7'>Google Form</a>
        """)
        mbox.setText(message)
        mbox.setWindowTitle('About/Info')
        mbox.setStandardButtons(QMessageBox.Ok)
        mbox.exec_()

    def displaySchema(self):
            # Setting the Row count of schema based on number of variables
            self.schemaTable.setRowCount(len(list(metaVarToLabels.keys())))

            # Displaying Variable Name and Variable Label
            row_number = 0
            for key, value in metaVarToLabels.items():
                self.schemaTable.setItem(row_number, 0, QTableWidgetItem(str(key)))
                if value != None:
                    self.schemaTable.setItem(row_number, 1, QTableWidgetItem(str(value)))
                row_number += 1

            # Displaying Variable Format and Type
            row_number = 0
            for value in metaVarToFormat.values():
                # Logic for type
                type = ''
                if value[0] == 'F':
                    type = 'Numeric'
                elif value[0] == 'A':
                    type = 'String'
                elif value[0] == 'D':
                    type = 'Date-Time'
                self.schemaTable.setItem(row_number, 2, QTableWidgetItem(str(type)))
                row_number += 1
            self.schemaTable.setEditTriggers(QAbstractItemView.NoEditTriggers)

    def displayStats(self):

        global df

        # Storing Variable Name and TypeName
        self.varName = self.schemaTable.item(self.schemaTable.currentRow(), 0).text()
        self.typeName = self.schemaTable.item(self.schemaTable.currentRow(), 2).text()

        # Clear all data from Frequency Table
        for i in reversed(range(self.freqTable.rowCount())):
            self.freqTable.removeRow(i)

        try:

            # Checks if Value Label is present
            valueLabelBool = bool(metaVarToValueLabels.get(self.varName, ''))

            # Dictionary for Count and Percentage
            freqCountDict = dict(df[self.varName].value_counts().sort_index())
            freqPercentageDict = dict(df[self.varName].value_counts(normalize=True).sort_index())

            if valueLabelBool and self.typeName == 'Numeric': # This is for Numeric Variables with value labels present

                # Dictionary for Value Label. Setting Number of Rows for Table
                valueLabelDict = metaVarToValueLabels.get(self.varName, '')
                self.freqTable.setRowCount(len(list(valueLabelDict.keys())) + 1)

                row_number = 0
                for valueStr in valueLabelDict.keys():

                    # value is finalized based on input type. We need this to input for get() in dictionary
                    if type(valueStr) == float:
                        value = float(valueStr)
                    elif type(valueStr) == int:
                        value = int(valueStr)
                    else:
                        value = valueStr

                    # Display value and labels
                    self.freqTable.setItem(row_number, 0, QTableWidgetItem(str(value)))
                    self.freqTable.setItem(row_number, 1, QTableWidgetItem(str(valueLabelDict.get(value, ''))))

                    # Display Counts and Percentages
                    count = freqCountDict.get(value, '0')
                    self.freqTable.setItem(row_number, 2, QTableWidgetItem(str(count)))
                    percentage = str(round(freqPercentageDict.get(value, 0) * 100, 2)) + ' %'
                    self.freqTable.setItem(row_number, 3, QTableWidgetItem(str(percentage)))
                    row_number += 1

            else:

                # Setting Number of Rows for Table
                self.freqTable.setRowCount(len(list(freqCountDict.keys())) + 1)

                # Finalizing variable type (Numeric or String) and assigning first_col (value or labels)
                if self.typeName == 'Numeric':
                    # Numeric Variable
                    firstColumn = 0
                elif self.typeName == 'String':
                    # String Variable
                    firstColumn = 1

                # Display Counts and Percentages
                row_number = 0
                sumCounter = 0
                for key, value in freqCountDict.items():
                    if key != '':
                        self.freqTable.setItem(row_number, firstColumn, QTableWidgetItem(str(key)))
                        self.freqTable.setItem(row_number, 2, QTableWidgetItem(str(value)))
                        percentage = str(round(freqPercentageDict.get(key, '') * 100, 2)) + ' %'
                        self.freqTable.setItem(row_number, 3, QTableWidgetItem(str(percentage)))
                        row_number += 1
                        sumCounter += value

            # Displaying total
            self.freqTable.setItem(row_number, 1, QTableWidgetItem('Total'))
            if self.typeName == 'Numeric':
                self.freqTable.setItem(row_number, 2, QTableWidgetItem(str(sum(list(freqCountDict.values())))))
            elif self.typeName == 'String':
                self.freqTable.setItem(row_number, 2, QTableWidgetItem(str(sumCounter)))
            row_number += 1

            # Displaying Statistics for Numeric Variables
            if self.typeName == 'Numeric':

                listStatsName = ['Minimum', 'Maximum', 'Mean', 'Median', 'Stddev', 'Variance']
                listStatsValue = []
                listStatsValue.append(df[self.varName].min())
                listStatsValue.append(df[self.varName].max())
                listStatsValue.append(round(df[self.varName].mean(), 2))
                listStatsValue.append(round(df[self.varName].median(), 2))
                listStatsValue.append(round(df[self.varName].std(), 2))
                listStatsValue.append(round(df[self.varName].var(), 2))

                # Blank Row
                self.freqTable.insertRow(row_number)
                row_number += 1

                for i in range(len(listStatsName)):
                    self.freqTable.insertRow(row_number)
                    self.freqTable.setItem(row_number, 1, QTableWidgetItem(str(listStatsName[i])))
                    self.freqTable.setItem(row_number, 2, QTableWidgetItem(str(listStatsValue[i])))
                    row_number += 1

        except:
            pass

    def searchTable(self):

        # Searches for variable name in Schema Table and highlights the row
        columnOfInterest = 0
        valueOfInterest = self.filterEntry.text()
        if valueOfInterest != '':
            self.filterEntry.setText('')
            self.schemaTable.setFocus()
            try:
                for rowIndex in range(self.schemaTable.rowCount()):
                    twItem = self.schemaTable.item(rowIndex, columnOfInterest)
                    if twItem.text() == valueOfInterest:
                        self.schemaTable.setCurrentCell(rowIndex, columnOfInterest)
            except:
                pass
        else:
            QMessageBox.information(self, 'Info', 'Search query cannot be empty')

# Creating Export Window
class Export(QWidget): # This is used to take inputs for export
    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon('icons/icon.png'))
        self.setGeometry(int(w/3), int(h/3), 200, 170)
        self.setFixedSize(self.size())
        self.UI()
        self.show()

    def UI(self):
        self.widgets()
        self.layouts()

    def widgets(self):
        self.excelValueExport = QRadioButton('Excel Export with Values')
        self.excelTextExport = QRadioButton('Excel Export with Text')
        self.excelTextExport.setChecked(True)
        self.optionDataMap = QCheckBox('Add DataMap')
        self.statusText = QLabel()
        self.confirmBtn = QPushButton('Submit')
        self.confirmBtn.clicked.connect(self.exportData)

    def layouts(self):
        self.mainLayout = QVBoxLayout()
        self.mainLayout.addWidget(self.excelValueExport)
        self.mainLayout.addWidget(self.excelTextExport)
        self.mainLayout.addWidget(self.optionDataMap)
        self.mainLayout.addStretch()
        self.mainLayout.addWidget(self.statusText)
        self.mainLayout.addStretch()
        self.mainLayout.addWidget(self.confirmBtn)
        self.setLayout(self.mainLayout)

    def exportData(self):

        directory = QFileDialog.getSaveFileName(self, 'Save File', '', 'Excel Files (*.xslx)')
        filename = directory[0]
        positiondot = filename.rfind('.')
        # This is used to ensure that extension xlsx is always present
        filename = filename[0:positiondot] + '.xlsx'

        # Data Export
        self.statusText.setText('Exporting Data ...')
        qApp.processEvents()
        if self.excelValueExport.isChecked():
            df.to_excel(filename, sheet_name='Data', index=False)
        elif self.excelTextExport.isChecked():
            dfText.to_excel(filename, sheet_name='Data', index=False)

        # Adding Data Map Sheet to exported file
        if self.optionDataMap.isChecked():
            self.statusText.setText('Generating Data Map ...')
            qApp.processEvents()
            workbook = openpyxl.load_workbook(filename)
            datamapSheet = workbook.create_sheet(index=1, title='Data Map')
            row_number = 1
            varToValueLabelDict = dict(metaVarToValueLabels)
            varToLabelDict = dict(metaVarToLabels)
            for x in varToLabelDict:
                # Print Variable Name and Label
                datamapSheet.cell(row=row_number, column=1).value = x
                datamapSheet.cell(row=row_number, column=2).value = varToLabelDict[x]
                row_number += 1
                # Print Value Label if present
                if bool(metaVarToValueLabels.get(x, '')):
                    for y in varToValueLabelDict[x]:
                        datamapSheet.cell(row=row_number, column=1).value = y
                        datamapSheet.cell(row=row_number, column=2).value = varToValueLabelDict[x][y]
                        row_number += 1
                row_number += 1
            workbook.save(filename)

        self.close()
        QMessageBox.information(self, 'Info', 'Data has been exported to \n%s' % filename)

def main():
    App = QApplication(sys.argv)
    window = Main()
    sys.exit(App.exec_())


if __name__ == '__main__':
    main()
